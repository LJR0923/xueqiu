import json
import os
from openpyxl import load_workbook
import pandas as pd
import datetime

def convert_timestamp_to_formatted_time(timestamp):
    timestamp_in_seconds = timestamp / 1000
    date_time = datetime.datetime.fromtimestamp(timestamp_in_seconds)
    formatted_date_time = date_time.strftime('%Y年%m月%d日 %H:%M:%S')
    return formatted_date_time

class XueqiucommontPipeline:
    def open_spider(self, spider):
        self.excel_path = 'comments.xlsx'
        self.data_frames = {
            ('first', 'sh_sz'): pd.DataFrame(),
            ('second', 'sh_sz'): pd.DataFrame(),
            ('third', 'sh_sz'): pd.DataFrame(),
            ('first', 'us'): pd.DataFrame(),
            ('second', 'us'): pd.DataFrame(),
            ('third', 'us'): pd.DataFrame(),
        }
        self.sheet_titles = {
            ('first', 'sh_sz'): ['股票代码', '评论ID', '用户ID', '是否是专栏', '评论内容', '图片链接', '创建时间',
                                 '回复数', '点赞数', '转发数', '目标', '标记', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname'],
            ('second', 'sh_sz'): ['股票代码', '评论ID', '根评论ID', '用户ID', '评论内容', '图片链接', '创建时间',
                                  '回复数', '点赞数', '标记', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname'],
            ('third', 'sh_sz'): ['股票代码', '评论ID', '根评论ID',  '评论内容', '图片链接', '创建时间',
                                 '回复数', '点赞数', '标记', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname'],
            ('first', 'us'): ['股票代码', '评论ID', '用户ID', '扩展信息', '评论内容', '图片链接', '创建时间', '回复数',
                              '点赞数', '转发数', '目标', '标记', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname'],
            ('second', 'us'): ['股票代码', '评论ID', '根评论ID', '用户ID', '评论内容', '图片链接', '创建时间', '回复数',
                               '点赞数', '标记', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname'],
            ('third', 'us'): ['股票代码', '评论ID', '根评论ID',  '评论内容', '图片链接', '创建时间', '回复数',
                              '点赞数', '标记', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname'],
        }

    def process_item(self, item, spider):
        item["data"]["user"]["user_blocking"] = item["data"]["user"].get("blocking","")
        item["data"]["user"]["user_description"] = item["data"]["user"].get("description","")
        item["data"]["user"]["user_donate_count"] = item["data"]["user"].get("donate_count","")
        item["data"]["user"]["user_id"] = item["data"]["user"].get("id","")
        item["data"]["user"]["user_created_at"] = item["data"]["user"].get("created_at","")
        del item["data"]["user"]["blocking"]
        del item["data"]["user"]["description"]
        del item["data"]["user"]["donate_count"]
        del item["data"]["user"]["id"]
        if 'created_at' in item["data"]["user"]:
            del item["data"]["user"]["created_at"]
        # 更新item内内容 user
        if 'user' in item['data']['user']:
            del item["data"]["user"]["user"]["description"]
            del item["data"]["user"]["user"]["blocking"]
            del item["data"]["user"]["user"]["id"]
            del item["data"]["user"]["user"]["donate_count"]
            item["data"]["user"]["user_created_at"] = item["data"]["user"]["user"].get("created_at", "")
            del item["data"]["user"]["user"]["created_at"]
            # 如果存在，更新 'data' 字典
            item["data"].update(item["data"]["user"]["user"])
        item['data'].update(item['data']['user'])


        item["data"]["created_at"] = convert_timestamp_to_formatted_time(item["data"]["created_at"])
        pic = item["data"].get("pic", "")
        item["data"]["pic"] = item["data"].get("pic", "").replace("thumb", "custom") if pic else ""

        key = (item['commont_floor'], item['Stock_type'])
        titles = self.sheet_titles[key]

        # 根据评论层级和股票类型选择不同的键
        if item['commont_floor'] == "first":
            keys = ['symbol', 'id', 'user_id', 'expend', 'description', 'pic', 'created_at', 'reply_count',
                    "like_count", 'retweet_count', "target", 'flags', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname']
        elif item['commont_floor'] == "second":
            keys = ['symbol', 'id', 'root_in_reply_to_status_id', 'user_id', 'description', 'pic', 'created_at',
                    'reply_count', "like_count", 'flags', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname']
        elif item['commont_floor'] == "third":
            keys = ['id', 'root_in_reply_to_status_id', 'user_id', 'description', 'pic', 'created_at', 'reply_count',
                    "like_count", 'flags', 'allow_all_stock','block_status','user_blocking','city','common_count','user_description',"user_donate_count",'followers_count','fortuneUser','friends_count','gender',"user_id",'province',"user_created_at",'screen_name','st_color','status','status_count','step','subscribeable','type','anonymous','areaCode','donate_snowcoin','truncated','verified','verified_description','verified_type','verified_realname']

        # 提取数据
        row_data = {title: item['data'].get(key, '') for title, key in zip(titles, keys)}
        row_df = pd.DataFrame([row_data])

        # 定义工作表名称
        sheet_name = f"{key[1]}{key[0]}级评论"

        try:
            # 检查 Excel 文件是否存在
            if os.path.exists(self.excel_path):
                # 加载现有工作簿
                with pd.ExcelWriter(self.excel_path, engine="openpyxl", mode="a", if_sheet_exists='overlay') as writer:
                    workbook = writer.book
                    # 检查 sheet 是否存在
                    if sheet_name in workbook.sheetnames:
                        # 如果 sheet 存在，追加数据
                        startrow = workbook[sheet_name].max_row
                        row_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                    else:
                        # 如果 sheet 不存在，创建新 sheet 并写入表头
                        row_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            else:
                # 如果文件不存在，创建新文件并写入表头
                with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                    row_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        except Exception as e:
            print(f"Error processing item: {e}")

        return item

    def close_spider(self, spider):
        # 在爬虫关闭时写入数据
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a') as writer:
            for key, df in self.data_frames.items():
                sheet_name = f"{key[1]}{key[0]}级评论"
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
